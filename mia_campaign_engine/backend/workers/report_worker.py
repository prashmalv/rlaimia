"""
Report Worker — generates a per-campaign Excel file after all jobs are complete.

Excel layout:
  One sheet per campaign with columns:
    S.No | Person Name | Age | Persona | Phase |
    Image URL | Video URL (compressed 5MB) | AI Avatar Video URL (compressed 5MB) |
    Status | Error

The compressed videos are generated on-demand here and stored at:
  reports/{campaign_id}/videos/{job_id}_compressed.mp4
  reports/{campaign_id}/videos/{job_id}_heygen_compressed.mp4

The Excel file is stored at:
  reports/{campaign_id}/campaign_report.xlsx

A 72-hour SAS URL is generated and saved on the Campaign record.
"""

import io
import logging
import tempfile
import os
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import sys
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
import config
from backend.app import azure_storage

logger = logging.getLogger(__name__)

TARGET_VIDEO_MB  = 5.0
TARGET_VIDEO_BYTES = int(TARGET_VIDEO_MB * 1024 * 1024)
REPORT_CONTAINER = config.AZURE_BLOB_CONTAINER_VID   # reuse video container for reports


def _get_video_url_for_report(
    original_blob_key: str,
    compressed_blob_key: str,
    container: str,
) -> Optional[str]:
    """
    Return a 30-day SAS URL for the video, compressing only if size > 5 MB.

    - If original ≤ 5 MB  → return fresh SAS of original (no re-upload needed)
    - If original > 5 MB  → compress, upload compressed copy, return its SAS
    - On any error        → return None
    """
    from backend.workers.video_worker import compress_video

    try:
        original_bytes = azure_storage.read_blob_bytes(original_blob_key, container)
    except Exception as e:
        logger.warning(f"[report] Could not download {original_blob_key}: {e}")
        return None

    size_mb = len(original_bytes) / 1024 / 1024

    if len(original_bytes) <= TARGET_VIDEO_BYTES:
        # Already within limit — just refresh SAS on the original blob
        logger.info(f"[report] {original_blob_key}: {size_mb:.1f} MB ≤ 5 MB, no compression needed")
        try:
            return azure_storage.get_sas_url(original_blob_key, container, hours=720)
        except Exception as e:
            logger.warning(f"[report] SAS refresh failed for {original_blob_key}: {e}")
            return None

    # Needs compression
    logger.info(f"[report] {original_blob_key}: {size_mb:.1f} MB > 5 MB, compressing…")
    try:
        compressed = compress_video(original_bytes, target_mb=TARGET_VIDEO_MB)
    except Exception as e:
        logger.warning(f"[report] Compression failed ({e}), using original")
        compressed = original_bytes

    try:
        azure_storage.upload_bytes(
            compressed, compressed_blob_key,
            container=REPORT_CONTAINER,
            content_type="video/mp4",
        )
        url = azure_storage.get_sas_url(compressed_blob_key, REPORT_CONTAINER, hours=720)
        logger.info(f"[report] Compressed: {size_mb:.1f} MB → {len(compressed)/1024/1024:.1f} MB")
        return url
    except Exception as e:
        logger.warning(f"[report] Upload of compressed video failed: {e}")
        return None


def generate_campaign_report(campaign_id: str) -> Optional[str]:
    """
    Build the Excel report for a completed campaign.
    Compresses all videos to ~5MB, uploads them, then writes the Excel.

    Returns the SAS URL of the uploaded Excel, or None on failure.
    Saves report_blob_key and report_url on the Campaign DB record.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # ── Load all jobs from DB (sync, called in asyncio.to_thread) ────────────
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker, Session
    from backend.app.models import Campaign, Job

    db_url = config.DATABASE_URL.replace("+asyncpg", "").replace("+aiosqlite", "")
    engine = create_engine(db_url, connect_args={"check_same_thread": False}
                           if "sqlite" in db_url else {})
    SessionLocal: type[Session] = sessionmaker(bind=engine)
    db = SessionLocal()

    try:
        campaign = db.query(Campaign).filter(Campaign.id == campaign_id).first()
        if not campaign:
            logger.error(f"[report] Campaign {campaign_id} not found")
            return None

        jobs = db.query(Job).filter(Job.campaign_id == campaign_id)\
                             .order_by(Job.created_at).all()
    finally:
        db.close()

    has_video  = campaign.generate_videos
    has_avatar = bool(campaign.heygen_talking_photo_id or campaign.heygen_video_template_id)

    # ── Build Excel workbook ──────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Campaign Report"

    # Header style
    header_fill   = PatternFill("solid", fgColor="1A1A2E")
    header_font   = Font(name="Calibri", bold=True, color="E8C97E", size=11)
    center_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align    = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin_border   = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    # ── Title row ──
    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = f"Mia Campaign Report — {campaign.name}"
    title_cell.font  = Font(name="Calibri", bold=True, size=14, color="1A1A2E")
    title_cell.fill  = PatternFill("solid", fgColor="E8C97E")
    title_cell.alignment = center_align
    ws.row_dimensions[1].height = 28

    # ── Meta row ──
    ws.merge_cells("A2:M2")
    meta_cell = ws["A2"]
    ist_time = datetime.now(timezone.utc).strftime("%d %b %Y %H:%M IST")
    meta_cell.value = (
        f"Campaign ID: {campaign_id}  |  "
        f"Total: {campaign.total_jobs}  |  "
        f"Status: {campaign.status}  |  "
        f"Generated: {ist_time}"
    )
    meta_cell.font = Font(name="Calibri", size=10, italic=True, color="666666")
    meta_cell.alignment = left_align
    ws.row_dimensions[2].height = 18

    # ── Column headers ──
    headers = [
        ("S.No",        8),
        ("Person Name", 22),
        ("Age",         7),
        ("Persona",     14),
        ("Phase",       10),
        ("Status",      11),
        ("Image",       18),
        ("Image URL",   70),
    ]
    if has_video:
        headers.append(("Video",    18))
        headers.append(("Video URL (≤5MB)", 70))
    if has_avatar:
        headers.append(("AI Avatar",    18))
        headers.append(("AI Avatar URL (≤5MB)", 70))
    headers.append(("Error", 30))

    for col_idx, (header, width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    # ── Data rows ──
    alt_fill = PatternFill("solid", fgColor="F5F5F5")

    for row_num, job in enumerate(jobs, start=1):
        row_idx = row_num + 3
        fill    = PatternFill("solid", fgColor="FFFFFF") if row_num % 2 else alt_fill

        # Compress and upload video if needed
        video_dl_url  : Optional[str] = None
        heygen_dl_url : Optional[str] = None

        if has_video and job.video_blob_key:
            compressed_key = f"reports/{campaign_id}/videos/{job.id}_compressed.mp4"
            video_dl_url = _get_video_url_for_report(
                job.video_blob_key, compressed_key,
                container=config.AZURE_BLOB_CONTAINER_VID,
            )

        if has_avatar and job.heygen_video_blob_key:
            compressed_key = f"reports/{campaign_id}/videos/{job.id}_heygen_compressed.mp4"
            heygen_dl_url = _get_video_url_for_report(
                job.heygen_video_blob_key, compressed_key,
                container=config.AZURE_BLOB_CONTAINER_VID,
            )

        # Build image URL (refresh SAS)
        image_url: Optional[str] = None
        if job.image_blob_key:
            try:
                image_url = azure_storage.get_sas_url(
                    job.image_blob_key, config.AZURE_BLOB_CONTAINER_IMG, hours=720
                )
            except Exception:
                image_url = job.image_url

        # Row values: for URL columns we write (link_label, raw_url) pairs
        row_data = [
            row_num,
            job.person_name or "",
            job.age or "",
            job.persona or "",
            job.phase or "",
            job.status or "",
            ("Download ↗", image_url) if image_url else ("", ""),      # Image link + raw URL
        ]
        if has_video:
            row_data.append(("Download ↗", video_dl_url) if video_dl_url else ("", ""))
        if has_avatar:
            row_data.append(("Download ↗", heygen_dl_url) if heygen_dl_url else ("", ""))
        row_data.append(job.error_msg or "")

        col_idx = 1
        for value in row_data:
            if isinstance(value, tuple):
                # URL pair: first cell = hyperlink label, second cell = raw URL
                label, url = value
                # Hyperlink cell
                cell = ws.cell(row=row_idx, column=col_idx, value=label)
                cell.fill      = fill
                cell.border    = thin_border
                cell.alignment = center_align
                if url:
                    cell.hyperlink = url
                    cell.font = Font(name="Calibri", size=10, color="0563C1", underline="single", bold=True)
                else:
                    cell.font = Font(name="Calibri", size=10, color="999999")
                col_idx += 1
                # Raw URL cell
                url_cell = ws.cell(row=row_idx, column=col_idx, value=url or "")
                url_cell.fill      = fill
                url_cell.border    = thin_border
                url_cell.alignment = left_align
                url_cell.font      = Font(name="Calibri", size=9, color="444444")
            else:
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill      = fill
                cell.border    = thin_border
                cell.alignment = left_align
                cell.font      = Font(name="Calibri", size=10)
            col_idx += 1

        ws.row_dimensions[row_idx].height = 18

    # ── Upload Excel to Azure ─────────────────────────────────────────────────
    excel_buf = io.BytesIO()
    wb.save(excel_buf)
    excel_bytes = excel_buf.getvalue()

    blob_key = f"reports/{campaign_id}/campaign_report.xlsx"
    try:
        azure_storage.upload_bytes(
            excel_bytes, blob_key,
            container=REPORT_CONTAINER,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        report_url = azure_storage.get_sas_url(blob_key, REPORT_CONTAINER, hours=720)
        logger.info(f"[report] Excel uploaded: {blob_key} ({len(excel_bytes):,} bytes)")
    except Exception as e:
        logger.error(f"[report] Excel upload failed: {e}")
        return None

    # ── Save report URL on Campaign record ───────────────────────────────────
    db2 = SessionLocal()
    try:
        from sqlalchemy import update as sa_update
        from backend.app.models import Campaign as CampaignModel
        db2.execute(
            sa_update(CampaignModel)
            .where(CampaignModel.id == campaign_id)
            .values(report_blob_key=blob_key, report_url=report_url)
        )
        db2.commit()
    except Exception as e:
        logger.warning(f"[report] Could not save report_url to DB: {e}")
    finally:
        db2.close()

    return report_url
